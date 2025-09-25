# coding=utf-8
import re
import os
from openpyxl import load_workbook
import jieba
import jieba.posseg as psg
from typing import List
from frontend.generate_lexicon import generate_lexicon
from frontend.symbols import lang_s2p
from tools.log import Log as logger
from frontend.zh_frontend import Frontend, INITIALS, TONES
from frontend.g2pw import G2PWOnnxConverter



class DialectFrontend(Frontend):
    def __init__(self, dialect_type: str, g2p_model="g2pW", vocab_phones=None, vocab_tones=None, use_rhy=False, 
                 dialect_g2pW_model_path: str=None,):
        super().__init__(g2p_model=g2p_model, vocab_phones=vocab_phones, vocab_tones=vocab_tones, use_rhy=use_rhy)
        self.dialect_type = dialect_type
        self.rule1_dict, self.rule2_dict, self.rule3_dict = self.get_rule_dict()
        self.pinyin2phone = generate_lexicon(with_tone=True, with_erhua=True)  #方言的儿化音比较多，所以with_erhua=True
        
        new_pinyin2phone, new_phone, new_initial = lang_s2p[dialect_type]
        self.pinyin2phone.update(new_pinyin2phone)  # 更新方言新的拼音映射
        self.vocab_phones = self.vocab_phones + new_phone   # 更新phone 集合
        self.initals = INITIALS + new_initial # 更新声母
        
        self.dialect2pt_tone = {"0": "0", "1": "1", "2": "2", "3": "3", "4": "4", "5": "5"}
        self.pt2dialect_tone = {v: k for k, v in self.dialect2pt_tone.items()}
        
        self.dialect_g2pW_model = None
        if dialect_g2pW_model_path is not None:
            self.dialect_g2pW_model = G2PWOnnxConverter(model_dir=dialect_g2pW_model_path, style="pinyin", enable_non_tradional_chinese=True)
            self.dialect_polyphonic_char = list(self.dialect_g2pW_model.polyphonic_chars_new)
          
    
        
    # 从文本和对应的拼音上获取phone
    def get_phone_from_ppinyin(self, text:str, ppinyin: str):
        # print(f"Use ref pinyin to get phone: {text}|{ppinyin}")
        word_pinyins_dict = {}
        text = text.replace("……", "…").replace("......", "…").replace("...", "…").replace("；", ",").replace("、", ",").replace("：", ",").replace("“", "").replace("”", "").replace("《", "").replace("》", "") # text
        ppinyin = ppinyin.replace("……", "…").replace("......", "…").replace("...", "…").replace("；", ",").replace("、", ",").replace("：", ",").replace("“ ", "").replace("” ", "").replace("《 ", "").replace("》 ", "").replace("“", "").replace("”", "").replace("《", "").replace("》", "") # pinyin
        ppinyin = re.sub(r'\s*#\s*', '#', ppinyin)
        text_parts = text.split("#")
        text_parts = [item for item in text_parts if item.strip()]
        ppinyin_parts = ppinyin.split("#")
        ppinyin_parts = [item for item in ppinyin_parts if item.strip()]
        assert len(text_parts) == len(ppinyin_parts), print(f"text_parts's length: {len(text_parts)} is not equal to ppinyin_parts's length: {len(ppinyin_parts)}")
        for i in range(len(text_parts)):
            text_part = text_parts[i]
            ppinyin_part = ppinyin_parts[i]
            # 方言的音调转普通话的音调
            ppinyin_part_list = [x if x[-1] not in self.dialect2pt_tone.keys() else x[:-1]+ self.dialect2pt_tone[x[-1]] 
                          for x in ppinyin_part.split(" ")]
            word_pinyins_dict[(text_part, i)] = ppinyin_part_list + ["None"] * (len(text_part)-len(ppinyin_part_list))
            assert len(text_part) == len(word_pinyins_dict[(text_part, i)]), print(f"text not equal to pinyin in {text_part} and {word_pinyins_dict[text_part]}")
        
        phones_list, word2ph, tones_list, _, oop = self.get_phone_from_word_pinyins_dict(word_pinyins_dict)

        return phones_list, word2ph, tones_list, ppinyin, oop

    
    def get_phone_from_word_pinyins_dict(self, word_pinyins_dict):
        initials = []
        finals = []
        oop = []
        for word, pos in word_pinyins_dict.keys():
            word_pinyins = word_pinyins_dict[(word, pos)]
            sub_initials = []
            sub_finals = []
            
            for pinyin, char in zip(word_pinyins, word):
                if pinyin is None or pinyin == "None":
                    pinyin = "None"

                pinyin = pinyin.replace("u:", "v")

                if pinyin in self.pinyin2phone:
                    initial_final_list = self.pinyin2phone[pinyin].split(" ")
                    if len(initial_final_list) == 2:
                        sub_initials.append(initial_final_list[0])
                        sub_finals.append(initial_final_list[1])
                    elif len(initial_final_list) == 1:
                        sub_initials.append("")
                        sub_finals.append(initial_final_list[1])
                else:
                    if pinyin not in self.punc and pinyin != "None":
                        oop.append(pinyin[:-1] if pinyin[-1].isdigit() else pinyin)
                        
                    sub_initials.append(pinyin)
                    sub_finals.append(pinyin)

            initials.append(sub_initials)
            finals.append(sub_finals)
        initials = sum(initials, [])
        finals = sum(finals, [])

        word2ph = []
        phones = []
        tones_list = []
        for c, v in zip(initials, finals):
            if (c=="None") or (c == v and c not in self.vocab_phones and c not in self.punc):
                new_word2ph = float(word2ph[-1]/2)
                word2ph[-1] -= new_word2ph
                word2ph.append(new_word2ph)
                
            if c and c not in self.vocab_phones and c not in self.punc:
                continue

            if c and c not in self.punc:
                phones.append(c)
                tones_list.append(int(v[-1]))
            if c and c in self.punc:
                phones.append(c)
                tones_list.append(0)

            if v and v not in self.punc and v not in self.rhy_phns:
                phones.append(v[:-1])
                tones_list.append(int(v[-1]))

            if c and v and c != v:
                word2ph.append(2)
            else:
                word2ph.append(1)
        oop = list(set(oop))
        return phones, word2ph, tones_list, "", oop
        
    
    
    def get_rule_dict(self):
        """
        生成三种规则的字典，并且根据rule3生成自定义词典文件
        rule1_file : A: <字>     E: <普通话调>    F: <河南话调>      需要去掉第一行
        rule2_file : A: <字>	B: <需要改的原始读音，空说明无论什么读音都改>		D: <修改后读音>    F:<一些备注>
        rule3_file : A: <字词>    B: <最终读音>        需要去掉第一行
        """
        rule1_file = f"frontend/dialect/{self.dialect_type}/rule1.xlsx"
        rule2_file = f"frontend/dialect/{self.dialect_type}/rule2.xlsx"
        rule3_file = f"frontend/dialect/{self.dialect_type}/rule3.xlsx"
        custom_words = []
        rule1_dict = {}
        if os.path.isfile(rule1_file):
            wb = load_workbook(rule1_file)
            sheet1 = wb['Sheet1']
            for row in sheet1.iter_rows(min_row=2, values_only=True):
                row_context = list(row)
                rule1_dict[row_context[0]+str(row_context[4])] = str(row_context[5])
        
        rule2_dict = {}
        sheet1 = load_workbook(rule2_file)['Sheet1']
        for row in sheet1.iter_rows(min_row=1, values_only=True):
            row_context = list(row)
            wword = row_context[0]
            pt_pinyin = row_context[1]
            dialect_pinyin = row_context[3]
            if wword not in rule2_dict.keys():
                rule2_dict[wword] = [[pt_pinyin], [dialect_pinyin]]
            else:
                rule2_dict[wword][0].append(pt_pinyin)
                rule2_dict[wword][1].append(dialect_pinyin)
            
        rule3_dict = {}
        sheet1 = load_workbook(rule3_file)['Sheet1']
        for row in sheet1.iter_rows(min_row=2, values_only=True):
            row_context = list(row)
            sword = row_context[0]
            if len(sword) == 1:  # 如果是单个字， 加入到rule2_dict 中
                if sword not in rule2_dict.keys():
                    rule2_dict[sword] = [[None], [row_context[1]]]
                else:
                    rule2_dict[sword][0].append(None)
                    rule2_dict[sword][1].append(row_context[1])
            else:
                rule3_dict[sword] = row_context[1]
                custom_words.append(sword)
        
        print(f"{self.dialect_type} custom words num: {len(custom_words)}")
        if custom_words != []:   
            custom_words_file = f"frontend/dialect/{self.dialect_type}/custom_words.txt"
            with open(custom_words_file, "w", encoding='utf-8') as f:
                for custom_word in custom_words:
                    f.write(f"{custom_word} 10000000000\n")      
            jieba.load_userdict(custom_words_file)  # 结巴分词加载自定义词典
         
        return rule1_dict, rule2_dict, rule3_dict
    

    
    def create_pinyin_dict(self, pairs, pinyins):
        pinyin_dict = {}
        st = 0

        for word, pos in pairs:
            end = st+len(word)
            pinyin_dict[word] = " ".join(pinyins[st: end])
            st = end
        
        return pinyin_dict
    
    def combine_pinyins(self, pinyins, dialect_pinyins, pos):
        # pos 中的index对应的位置选择 dialect_pinyins， 反之选择 pinyins
        result = []
        for i in range(len(pinyins)):
            if i in pos:
                result.append(dialect_pinyins[i])
            else:
                result.append(pinyins[i])
        return result
    
    def get_polyphone_pos(self, seg):
        pos = []
        for i in range(len(seg)):
            word = seg[i]
            if word in self.dialect_polyphonic_char:
                pos.append(i)
        return pos
          

    
    # 普通话读音到方言的读音基本流程, rule1 -> rule2 -> 方言多音字 -> 获取词(结巴分词)读音 -> rule3
    def pt_to_dialect_base(self, seg, seg_cut, pinyins):
        # seg 是一句话， pinyins 是这句话对应的拼音，按道理每个字都有一个对应的拼音。
        # print(seg, pinyins)
        assert len(seg) == len(pinyins), print("There was an error when getting the pinyin")  
        for i in range(len(seg)):   # 这里是对单个字读音的变换，过规则1和规则2
            word_tone = seg[i] + pinyins[i][-1]
            if word_tone in self.rule1_dict.keys():
                pinyins[i] = pinyins[i][:-1] + self.rule1_dict[word_tone]
            
            if seg[i] in self.rule2_dict.keys():
                print("rule2 self.rule2_dict[seg[i]]: ", self.rule2_dict[seg[i]])
                if pinyins[i] in self.rule2_dict[seg[i]][0]:
                    index = self.rule2_dict[seg[i]][0].index(pinyins[i])
                    pinyins[i] = self.rule2_dict[seg[i]][1][index]
                elif None in self.rule2_dict[seg[i]][0]:
                    index = self.rule2_dict[seg[i]][0].index(None)
                    pinyins[i] = self.rule2_dict[seg[i]][1][index]
                
        # 方言多音字消歧
        ## 首先获取这个方言独有的多音字对应的字: self.dialect_polyphonic_char
        ## 然后对这句话多音字对应的位置进行记录
        if self.dialect_g2pW_model is not None:
            pos = self.get_polyphone_pos(seg)
            if pos != []:
                try:
                    ## 然后对这句话经过方言多音字进行整句话的拼音获取
                    dialect_pinyins = self.dialect_g2pW_model(seg)[0]  # 这里好像有失败的情况， 嗯 这种字眼，后续测试看看，应该是pypinyin 过不了导致的
                    ## 结合过方言多音字之后的拼音和之前的拼音，多音字对应的位置进行拼音修改，其他的保持原来的不变
                    assert len(dialect_pinyins) == len(pinyins), print(f"Error in getting the pinyin of a dialect polyphone: dialect pinyin length: {len(dialect_pinyins)} not equal to raw pinyin length: {len(pinyins)}.")
                    pinyins = self.combine_pinyins(pinyins, dialect_pinyins, pos)
                except Exception as e:
                    logger.warning("g2pW_model error. [%s] not in g2pW dict, use g2pM" % seg)
            
             
        # 获取分词后的分词以及对应的拼音
        word_pinyin = self.create_pinyin_dict(seg_cut, pinyins)
        for word in word_pinyin.keys():
            if word in self.rule3_dict.keys():
                word_pinyin[word] = self.rule3_dict[word]
        
        return word_pinyin
    
    def change_to_pttone(self, word_pinyin):
        for word in word_pinyin.keys():
            ppinyin = word_pinyin[word]
            new_pinyin = [x if x[-1] not in self.dialect2pt_tone.keys() else x[:-1]+ self.dialect2pt_tone[x[-1]] 
                            for x in ppinyin.split(" ")]
            word_pinyin[word] = new_pinyin + ["None"] * (len(word) - len(new_pinyin))   # 如果有连读的情况，补 "None" 让每个字都有对应的 pinyin
        return word_pinyin
        
    
    def pt_to_dialect(self, seg, seg_cut, pinyins):
        word_pinyin = self.pt_to_dialect_base(seg, seg_cut, pinyins)
        word_pinyin = self.change_to_pttone(word_pinyin)
        return word_pinyin
    
    
    def get_splited_phonemes_tones(
        self, sentences: List[str], merge_sentences: bool = True, with_erhua: bool = True, dialect_tone: bool=False
    ) -> List[List[str]]:
        phones_list = []
        tones_list = []
        word2ph = []
        ppinyins_list = []
        oop = []

        for seg in sentences:
            seg_cut = psg.lcut(seg)  # 分词，生成词和词性， 例如：[pair('我', 'r'), pair('是', 'v'), pair('中国', 'ns'), pair('人', 'n'), pair(',', 'x')]
            try:
                pinyins = self.g2pW_model(seg)[0]
            except Exception:
                logger.warning(f"g2pW_model error. [{seg}] not in g2pW dict, use g2pM")
                pinyins = self.g2pM_model(seg, tone=True, char_split=False)
            # 这里获取的 pinyins 是每个字每个字的拼音， 例如：['wo3', 'shi4', 'zhong1', 'guo2', 'ren2', ',']
            
            # 普通话读音到方言读音的映射，最后转到普通话的音调
            word_pinyins_dict = self.pt_to_dialect(seg, seg_cut, pinyins)
               
            phones = []
            initials = []
            finals = []
            ppinyins = []
            
            for word, pos in seg_cut:
                sub_initials = []
                sub_finals = []
                sub_pinyins = []

                if pos == "eng":
                    continue

                word_pinyins = word_pinyins_dict[word]
                
                for pinyin, char in zip(word_pinyins, word):
                    if pinyin is None or pinyin == "None":
                        pinyin = "None"

                    pinyin = pinyin.replace("u:", "v")
                    sub_pinyins.append(pinyin)

                    if pinyin in self.pinyin2phone:
                        initial_final_list = self.pinyin2phone[pinyin].split(" ")
                        if len(initial_final_list) == 2:
                            sub_initials.append(initial_final_list[0])
                            sub_finals.append(initial_final_list[1])
                        elif len(initial_final_list) == 1:
                            sub_initials.append("")
                            sub_finals.append(initial_final_list[1])
                    else:
                        if pinyin not in self.punc and pinyin != "None":
                            oop.append(pinyin)
                        sub_initials.append(pinyin)
                        sub_finals.append(pinyin)

                if with_erhua:
                    sub_initials, sub_finals = self._merge_erhua(sub_initials, sub_finals, word, pos)
                initials.append(sub_initials)
                finals.append(sub_finals)
                
                for i in range(len(sub_finals)):
                    if sub_finals[i] != "None":
                        if sub_finals[i] not in self.punc:
                            sub_pinyins[i] = sub_pinyins[i][:-1] + sub_finals[i][-1]
                        # else:
                        #     ppinyins_list.append(finals[i])
                ppinyins.append(sub_pinyins)

            initials = sum(initials, [])
            finals = sum(finals, [])
            ppinyins = sum(ppinyins, [])
    

            for c, v in zip(initials, finals):
                if (c=="None") or (c == v and c not in self.vocab_phones and c not in self.punc):
                    new_word2ph = float(word2ph[-1]/2)
                    word2ph[-1] -= new_word2ph
                    word2ph.append(new_word2ph)
                    
                if c and c not in self.vocab_phones and c not in self.punc:
                    continue

                if c and c not in self.punc:
                    phones.append(c)
                    tones_list.append(int(v[-1]))
                if c and c in self.punc:
                    phones.append(c)
                    tones_list.append(0)

                if v and v not in self.punc and v not in self.rhy_phns:
                    phones.append(v[:-1])
                    tones_list.append(int(v[-1]))

                if c and v and c != v:
                    word2ph.append(2)
                else:
                    word2ph.append(1)

            phones_list.append(phones)
            ppinyins_list.append(ppinyins)

        if merge_sentences:
            merge_list = sum(phones_list, [])
            phones_list = []
            phones_list += merge_list
            
            merge_ppinyin_list = sum(ppinyins_list, [])
            ppinyins_list = []
            ppinyins_list += merge_ppinyin_list
            
        if dialect_tone:
            ppinyins_list_dialect_tone = ppinyins_list = [(item[:-1] + str(self.pt2dialect_tone[item[-1]])) if item[-1] in self.pt2dialect_tone.keys() else item for item in ppinyins_list]
            output_pinyin = " ".join(ppinyins_list_dialect_tone)
            logger.debug(f"dialect tone pinyin: {output_pinyin}")
        else:
            output_pinyin = ' '.join(ppinyins_list)
            logger.debug(f"putonghua pinyin: {' '.join(ppinyins_list)}")
            
        oop = list(set(oop))

        return phones_list, word2ph, tones_list, output_pinyin, oop
